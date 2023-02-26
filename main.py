import sys
import platform
from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (QCoreApplication, QPropertyAnimation, QDate, QDateTime, QMetaObject,
                            QObject, QPoint, QRect, QSize, QTime, QUrl, Qt, QEvent)
from PySide2.QtGui import (QBrush, QColor, QConicalGradient, QCursor, QFont,
                           QFontDatabase, QIcon, QKeySequence, QLinearGradient, QPalette, QPainter,
                           QPixmap, QRadialGradient)
from PySide2.QtWidgets import *

from openpyxl import load_workbook

from datetime import datetime

import time

import serial

import sqlite3

## ==> MAIN WINDOW
from ui_pages import Ui_MainWindow

workbook1 = load_workbook(filename="./database/crops.xlsx")

spreadsheet1 = workbook1.active
crops = spreadsheet1.max_row

workbook2 = load_workbook(filename='./database/et0.xlsx')

spreadsheet2 = workbook2.active

soil_types = ['Clay', 'Sand', 'Silt', 'Loam']


# cities = spreadsheet2["A"].value

# print(spreadsheet2.cell(row=spreadsheet2.max_row, column=1).value)

class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        ## EXIT BUTTON

        self.ui.closeButton.clicked.connect(lambda: self.close())

        # minimize

        self.ui.minimize_btn.clicked.connect(lambda: self.showMinimized())

        # crops comboBox
        for crop in range(2, spreadsheet1.max_row + 1):
            print(spreadsheet1.cell(row=crop, column=1).value)
            self.ui.comboBox.addItem(spreadsheet1.cell(row=crop, column=1).value)

        # cities comboBox

        for city in range(2, spreadsheet2.max_row + 1):
            print(spreadsheet2.cell(row=city, column=1).value)
            self.ui.comboBox_4.addItem(spreadsheet2.cell(row=city, column=1).value)

        # SOIL COMBOBOX

        for soil in soil_types:
            self.ui.comboBox_3.addItem(soil)

        # add crop button

        self.ui.pushButton.clicked.connect(lambda: self.add_crop())

        # Current Date as defualt date

        self.ui.dateEdit.setDateTime(QtCore.QDateTime.currentDateTime())

        self.ui.pushButton_2.clicked.connect(lambda: self.anim())

        self.ui.pushButton.clicked.connect(lambda: self.anim())

        def moveWindow(e):
            # Detect if the window is  normal size
            # ###############################################  

            if e.buttons() == Qt.LeftButton:
                # Move window
                self.move(self.pos() + e.globalPos() - self.clickPosition)
                self.clickPosition = e.globalPos()
                e.accept()

        self.ui.label_24.mouseMoveEvent = moveWindow

        # WIDGET TO MOVE

    def mousePressEvent(self, event):
        # ###############################################
        # Get the current position of the mouse
        self.clickPosition = event.globalPos()

    def anim(self):

        self.ui.frame_17.setHidden(True)

        self.card_1 = QPropertyAnimation(self.ui.frame_4, b'geometry')

        self.card_1.setDuration(500)
        self.card_1.setStartValue(QRect(110, 160, 281, 161))
        self.card_1.setEndValue(QRect(110, 210, 281, 161))

        self.card_1.start()

        self.card_2 = QPropertyAnimation(self.ui.frame_5, b'geometry')

        self.card_2.setDuration(500)
        self.card_2.setStartValue(QRect(411, 160, 281, 161))
        self.card_2.setEndValue(QRect(411, 210, 281, 161))

        self.card_2.start()

        self.card_3 = QPropertyAnimation(self.ui.frame_6, b'geometry')

        self.card_3.setDuration(500)
        self.card_3.setStartValue(QRect(712, 160, 281, 161))
        self.card_3.setEndValue(QRect(712, 210, 281, 161))

        self.card_3.start()

        self.card_4 = QPropertyAnimation(self.ui.frame_7, b'geometry')

        self.card_4.setDuration(500)
        self.card_4.setStartValue(QRect(1012, 160, 281, 161))
        self.card_4.setEndValue(QRect(1012, 210, 281, 161))

        self.card_4.start()

        self.card_5 = QPropertyAnimation(self.ui.frame_8, b'geometry')

        self.card_5.setDuration(500)
        self.card_5.setStartValue(QRect(110, 360, 581, 291))
        self.card_5.setEndValue(QRect(110, 410, 581, 291))

        self.card_5.start()

        self.card_6 = QPropertyAnimation(self.ui.frame_9, b'geometry')

        self.card_6.setDuration(500)
        self.card_6.setStartValue(QRect(710, 360, 581, 291))
        self.card_6.setEndValue(QRect(710, 410, 581, 291))

        self.card_6.start()

        self.ui.frame_17.show()

        # time.sleep(2)

        # self.ui.frame_17.hide()

        # self.card_1 = QPropertyAnimation(self.ui.frame_4, b'geometry')

        # self.card_1.setDuration(500)
        # self.card_1.setStartValue(QRect(110, 210, 281, 161))
        # self.card_1.setEndValue(QRect(110, 160, 281, 161))

        # self.card_1.start()

        # self.card_2 = QPropertyAnimation(self.ui.frame_5, b'geometry')

        # self.card_2.setDuration(500)
        # self.card_2.setStartValue(QRect(411, 160, 281, 161))
        # self.card_2.setEndValue(QRect(411, 210, 281, 161))

        # self.card_2.start()

        # self.card_3 = QPropertyAnimation(self.ui.frame_6, b'geometry')

        # self.card_3.setDuration(500)
        # self.card_3.setStartValue(QRect(712, 160, 281, 161))
        # self.card_3.setEndValue(QRect(712, 210, 281, 161))

        # self.card_3.start()

        # self.card_4 = QPropertyAnimation(self.ui.frame_7, b'geometry')

        # self.card_4.setDuration(500)
        # self.card_4.setStartValue(QRect(1012, 160, 281, 161))
        # self.card_4.setEndValue(QRect(1012, 210, 281, 161))

        # self.card_4.start()

        # self.card_5 = QPropertyAnimation(self.ui.frame_8, b'geometry')

        # self.card_5.setDuration(500)
        # self.card_5.setStartValue(QRect(110, 360, 581, 291))
        # self.card_5.setEndValue(QRect(110, 410, 581, 291))

        # self.card_5.start()

        # self.card_6 = QPropertyAnimation(self.ui.frame_9, b'geometry')

        # self.card_6.setDuration(500)
        # self.card_6.setStartValue(QRect(710, 360, 581, 291))
        # self.card_6.setEndValue(QRect(710, 410, 581, 291))

        # self.card_6.start()

        # self.notification = QPropertyAnimation(self.ui.frame_17, b'geometry')

        # self.notification.setDuration(500)
        # self.notification.setStartValue(QRect(61, -31, 821, 61))
        # self.notification.setEndValue(QRect(320, 140, 651, 41))

        # self.notification.start()

    def add_crop(self):
        crop_selected = self.ui.comboBox.currentText()
        date_selected = self.ui.dateEdit.date().toString("dd/MM/yyyy")
        city_selected = self.ui.comboBox_4.currentText()

        month = self.ui.dateEdit.date().toString("MM")
        day = self.ui.dateEdit.date().toString("dd")
        year = self.ui.dateEdit.date().toString("yyyy")

        if month == '01':
            month_name = 'Jan'
        if month == '02':
            month_name = 'Feb'
        if month == '03':
            month_name = 'Mar'
        if month == '04':
            month_name = 'Apr'
        if month == '05':
            month_name = 'May'
        if month == '06':
            month_name = 'Jun'
        if month == '07':
            month_name = 'Jul'
        if month == '08':
            month_name = 'Aug'
        if month == '09':
            month_name = 'Sep'
        if month == '10':
            month_name = 'Oct'
        if month == '11':
            month_name = 'Nov'
        if month == '12':
            month_name = 'Dec'

        current_mon = int(datetime.today().strftime('%m'))
        current_day = int(datetime.today().strftime('%d'))
        current_year = int(datetime.today().strftime('%Y'))

        # if int(year) > current_year or int(month) > current_mon and :
        #     if int(month) > current_mon:
        #         if 

        if date_selected > datetime.today().strftime('%d/%m/%Y'):
            self.ui.label_25.setStyleSheet(u"\n"
                                           "color:rgb(255, 102, 82); ")
            self.ui.label_25.setText('Error: Entered Date is greater than the current date.')
        else:
            self.ui.label_25.setStyleSheet(u"\n"
                                           "color:rgb(34, 182, 49); ")

            self.ui.label_25.setText('Crop successfully entered.')

        for month_ in range(2, spreadsheet2.max_column + 1):

            print(spreadsheet2.cell(row=1, column=month_).value)
            if month_name == spreadsheet2.cell(row=1, column=month_).value:
                et0_col = month_

        for city_ in range(2, spreadsheet2.max_row + 1):

            print(spreadsheet2.cell(row=city_, column=1).value)
            if city_selected == spreadsheet2.cell(row=city_, column=1).value:
                et0_row = city_

        for crop_ in range(2, spreadsheet1.max_row + 1):

            print(spreadsheet1.cell(row=city_, column=1).value)
            if crop_selected == spreadsheet1.cell(row=crop_, column=1).value:
                kc_row = crop_

        kc = spreadsheet1.cell(row=kc_row, column=et0_col).value
        eto = spreadsheet2.cell(row=et0_row, column=et0_col).value
        print(eto, kc)
        self.ui.label_5.setText(str("{:.2f}".format(eto * kc)))

        print(crop_selected)
        print(date_selected)
        print(city_selected)

        # Create a empty row at bottom of table
        numRows = self.ui.tableWidget.rowCount()
        self.ui.tableWidget.insertRow(numRows)
        # Add text to the row
        self.ui.tableWidget.setItem(numRows, 0, QtWidgets.QTableWidgetItem(crop_selected))
        self.ui.tableWidget.setItem(numRows, 1, QtWidgets.QTableWidgetItem(date_selected))
        self.ui.tableWidget.setItem(numRows, 2, QtWidgets.QTableWidgetItem(city_selected))
        # self.ui.tableWidget.setItem(numRows, 3, QtWidgets.QTableWidgetItem("Initial"))

        # Row Delete button


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
